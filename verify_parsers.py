"""Verification script — reads real XLS/XLSM files and dumps cell values
to verify parser column/row mappings.

Checks:
  WARN 1: Page 2 Plan 15 column (is it col 30 or 31?)
  WARN 2: Page 1 zone rows (31-38 vs 33-38)
  BUG 3:  Page 3 plan 16-30 column mapping
  BUG 1:  Master List XLSM — FDOT phase columns
"""

import json
import sys
from pathlib import Path

# ── BiTrans XLS verification ──────────────────────────────────────────

XLS_PATH = Path("C:/enock/fastapi/2401_NW 12 Av&NW 20 St_Complete.xls")


def verify_bitrans():
    import xlrd

    if not XLS_PATH.exists():
        print(f"[SKIP] BiTrans XLS not found: {XLS_PATH}")
        return

    wb = xlrd.open_workbook(str(XLS_PATH))
    print(f"Sheets: {wb.sheet_names()}\n")

    # ── WARN 2: Page 1 zone rows ──
    print("=" * 60)
    print("WARN 2: Page 1 zone rows (checking rows 30-39)")
    print("=" * 60)
    sheet1 = wb.sheet_by_name("Page 1 of 8")
    print(f"  Sheet dimensions: {sheet1.nrows} rows x {sheet1.ncols} cols")
    for r in range(30, min(40, sheet1.nrows)):
        vals = []
        for c in range(min(5, sheet1.ncols)):
            try:
                v = sheet1.cell_value(r, c)
            except IndexError:
                v = ""
            vals.append(repr(v))
        print(f"  Row {r}: {', '.join(vals)}")

    # ── WARN 1: Page 2 Plan 15 column ──
    print("\n" + "=" * 60)
    print("WARN 1: Page 2 — Coordination plans column verification")
    print("=" * 60)
    sheet2 = wb.sheet_by_name("Page 2 of 8")
    print(f"  Sheet dimensions: {sheet2.nrows} rows x {sheet2.ncols} cols")

    # Check cycle length row (row 26) across columns to find plan headers
    print("\n  Row 26 (cycle length) — scanning cols 0-35:")
    for c in range(min(36, sheet2.ncols)):
        v = sheet2.cell_value(26, c)
        if v:
            print(f"    col {c}: {v}")

    # Check Plan 14 and 15 specifically
    print("\n  Plans 13-15 detail (rows 26-38):")
    for c in [27, 29, 30, 31]:
        print(f"\n  --- Col {c} ---")
        for r in range(26, min(39, sheet2.nrows)):
            try:
                v = sheet2.cell_value(r, c)
            except IndexError:
                v = ""
            if v:
                print(f"    Row {r}: {v}")

    # Verify all 15 plans: dump plan headers from row 25 or a label row
    print("\n  Row 25 (possible plan labels):")
    for c in range(min(36, sheet2.ncols)):
        v = sheet2.cell_value(25, c) if 25 < sheet2.nrows else ""
        if v:
            print(f"    col {c}: {v}")

    # ── BUG 3: Page 3 plan 16-30 columns ──
    print("\n" + "=" * 60)
    print("BUG 3: Page 3 — Plans 16-30 column verification")
    print("=" * 60)
    sheet3 = wb.sheet_by_name("Page 3 of 8")
    print(f"  Sheet dimensions: {sheet3.nrows} rows x {sheet3.ncols} cols")

    # Scan row 4 (cycle length) for non-zero values
    print("\n  Row 4 (cycle length) — scanning all cols:")
    for c in range(min(25, sheet3.ncols)):
        v = sheet3.cell_value(4, c)
        if v:
            print(f"    col {c}: {v}")

    # Row 3 — possible plan number labels
    print("\n  Row 3 (possible labels) — scanning all cols:")
    for c in range(min(25, sheet3.ncols)):
        v = sheet3.cell_value(3, c)
        if v:
            print(f"    col {c}: {v}")

    # Row 2 — another possible label row
    print("\n  Row 2 (possible labels) — scanning all cols:")
    for c in range(min(25, sheet3.ncols)):
        v = sheet3.cell_value(2, c)
        if v:
            print(f"    col {c}: {v}")

    # Dump full structure for rows 0-17
    print("\n  Full dump rows 0-17:")
    for r in range(min(18, sheet3.nrows)):
        vals = []
        for c in range(min(25, sheet3.ncols)):
            v = sheet3.cell_value(r, c)
            if v:
                vals.append(f"c{c}={v}")
        if vals:
            print(f"    Row {r}: {', '.join(vals)}")

    # xlrd Book has no close() method


# ── Master List XLSM verification ────────────────────────────────────

XLSM_PATH = Path("C:/enock/fastapi/MDC Data Prep - Ver 4.4.xlsm")


def verify_master_list():
    import openpyxl

    if not XLSM_PATH.exists():
        print(f"[SKIP] Master List XLSM not found: {XLSM_PATH}")
        return

    print("\n" + "=" * 60)
    print("BUG 1: Master List XLSM — FDOT phase columns")
    print("=" * 60)

    wb = openpyxl.load_workbook(str(XLSM_PATH), read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")

    # Find the Master Intersection List sheet
    target = None
    for name in wb.sheetnames:
        if "master" in name.lower() and "list" in name.lower():
            target = wb[name]
            print(f"  Found sheet: '{name}'")
            break

    if not target:
        # Try all sheets
        for name in wb.sheetnames:
            if "list" in name.lower() or "intersection" in name.lower():
                target = wb[name]
                print(f"  Found sheet: '{name}'")
                break

    if not target:
        print("  [!] No master list sheet found. Listing all sheets:")
        for name in wb.sheetnames:
            s = wb[name]
            print(f"    '{name}': {s.max_row} rows x {s.max_column} cols")
        wb.close()
        return

    print(f"  Dimensions: {target.max_row} rows x {target.max_column} cols")

    # Dump first 6 rows (headers area)
    print("\n  Header area (rows 1-6):")
    rows = list(target.iter_rows(min_row=1, max_row=6, values_only=False))
    for row in rows:
        vals = []
        for cell in row:
            if cell.value is not None:
                vals.append(f"{cell.column_letter}{cell.row}={cell.value}")
        if vals:
            print(f"    {', '.join(vals)}")

    # Dump header row with column indices
    print("\n  Column headers (with indices):")
    header_rows = list(target.iter_rows(min_row=1, max_row=5, values_only=False))
    for row in header_rows:
        has_content = False
        for cell in row:
            if cell.value and any(kw in str(cell.value).lower()
                                  for kw in ["phase", "yellow", "red", "ped", "speed",
                                             "grade", "width", "cross", "asset", "polygon",
                                             "clear", "number", "location"]):
                has_content = True
                break
        if has_content:
            print(f"    Row {row[0].row}:")
            for cell in row:
                if cell.value is not None:
                    print(f"      Col {cell.column - 1} ({cell.column_letter}): {cell.value}")

    # Find a specific asset and dump all columns for it
    print("\n  Looking for asset '2401' in first column data...")
    found = False
    for row in target.iter_rows(min_row=2, max_row=min(target.max_row, 3000), values_only=False):
        cell_val = row[1].value if len(row) > 1 else None  # Column B
        if cell_val is None:
            cell_val = row[0].value  # Try column A
        if cell_val is not None:
            val_str = str(int(cell_val)) if isinstance(cell_val, float) else str(cell_val)
            if val_str.strip() == "2401":
                found = True
                print(f"  Found asset 2401 at row {row[0].row}:")
                for cell in row:
                    if cell.value is not None:
                        print(f"    Col {cell.column - 1} ({cell.column_letter}): {cell.value}")
                break

    if not found:
        # Show first few data rows to understand structure
        print("  Asset 2401 not found. Showing first 3 data rows:")
        data_rows = list(target.iter_rows(min_row=5, max_row=8, values_only=False))
        for row in data_rows:
            vals = []
            for cell in row:
                if cell.value is not None:
                    vals.append(f"c{cell.column - 1}({cell.column_letter})={cell.value}")
            if vals:
                print(f"    Row {row[0].row}: {', '.join(vals)}")

    wb.close()


if __name__ == "__main__":
    print("BiTrans → SEPAC Parser Verification Script")
    print("=" * 60)
    verify_bitrans()
    verify_master_list()
    print("\n\nDone.")
