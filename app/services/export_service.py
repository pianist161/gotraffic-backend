"""Export service — generate Excel, JSON, and PDF output for converted intersections."""

import io
import json
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session

from app.models.intersection import Intersection
from app.models.timing import PhaseTiming, CoordinationPlan, FdotOverride
from app.services.conversion_engine import compute_min_splits, compute_min_splits_without_peds, validate_splits


def export_intersection_json(db: Session, asset_number: str) -> dict | None:
    """Export intersection data as JSON."""
    intersection = db.query(Intersection).filter(
        Intersection.asset_number == asset_number
    ).first()
    if not intersection:
        return None

    timings = db.query(PhaseTiming).filter(
        PhaseTiming.intersection_id == intersection.id
    ).all()

    plans = db.query(CoordinationPlan).filter(
        CoordinationPlan.intersection_id == intersection.id
    ).order_by(CoordinationPlan.plan_number).all()

    return {
        "asset_number": intersection.asset_number,
        "location": intersection.location_name,
        "section": intersection.section,
        "equipment_type": intersection.equipment_type,
        "phase_timings": [
            {
                "bank": t.bank, "phase": t.phase_number,
                "ped_walk": t.ped_walk, "ped_fdw": t.ped_fdw,
                "min_green": t.min_green, "yellow_change": t.yellow_change,
                "red_clear": t.red_clear,
            }
            for t in timings
        ],
        "coordination_plans": [
            {
                "plan": p.plan_number, "cycle_length": p.cycle_length,
                "offset": p.offset,
                "force_offs": {
                    f"ph{i}": getattr(p, f"phase{i}_force_off")
                    for i in range(1, 9)
                },
                "sepac_splits": {
                    f"ph{i}": getattr(p, f"sepac_split{i}")
                    for i in range(1, 9)
                },
                "converted": p.converted,
            }
            for p in plans if p.cycle_length > 0
        ],
    }


def _compute_min_split_with_peds(timing: PhaseTiming) -> float:
    """VBA: MAX(MinGreen, Walk+FDW) + Yellow + RedClear + 1."""
    ped_total = timing.ped_walk + timing.ped_fdw
    green_component = max(timing.min_green, ped_total)
    return round(green_component + timing.yellow_change + timing.red_clear + 1, 1)


def _compute_min_split_without_peds(timing: PhaseTiming) -> float:
    """VBA: MinGreen + Yellow + RedClear + 1."""
    return round(timing.min_green + timing.yellow_change + timing.red_clear + 1, 1)


def export_intersection_excel(db: Session, asset_number: str) -> io.BytesIO | None:
    """Export intersection data as Excel workbook matching VBA 'Calculated Split Times' layout.

    Layout mirrors MDC Data Prep macro output:
    - Plans arranged in groups of 10 side-by-side
    - Each plan has BiTrans (ForceOff) and SEPAC (split) columns
    - Min split tables on right side (with/without PEDs, 3 banks)
    - Location/Asset info in rows 11-12
    """
    intersection = db.query(Intersection).filter(
        Intersection.asset_number == asset_number
    ).first()
    if not intersection:
        return None

    # Get all timings (all banks) for min-split tables
    all_timings = db.query(PhaseTiming).filter(
        PhaseTiming.intersection_id == intersection.id,
    ).order_by(PhaseTiming.bank, PhaseTiming.phase_number).all()

    # Build plan lookup: plan_number -> CoordinationPlan (all 30 slots)
    all_plans = db.query(CoordinationPlan).filter(
        CoordinationPlan.intersection_id == intersection.id,
    ).order_by(CoordinationPlan.plan_number).all()
    plan_map = {p.plan_number: p for p in all_plans}

    wb = Workbook()

    # --- Sheet 1: Calculated Split Times (VBA layout) ---
    ws = wb.active
    ws.title = "Calculated Split Times"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    sepac_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red for failed splits
    fail_font = Font(bold=True, color="9C0006")  # dark red text
    label_font = Font(bold=True)

    # Row 1: Title
    ws["A1"] = "Pattern data conversion from BiTrans forceoff point to SEPAC split times"
    ws["A1"].font = Font(bold=True, size=11)

    # --- Min Splits section (right side, cols M-W) ---
    # Organize timings by bank
    timings_by_bank = {}
    for t in all_timings:
        timings_by_bank.setdefault(t.bank, {})[t.phase_number] = t

    # "SEPAC Minimum Splits without PEDs" — M2
    ws.cell(row=2, column=13, value="SEPAC Minimum Splits without PEDs").font = label_font
    # "SEPAC Minimum Splits with PEDs" — R2
    ws.cell(row=2, column=18, value="SEPAC Minimum Splits with PEDs").font = label_font

    # Bank headers — row 3
    for offset, bank_num in [(0, 1), (1, 2), (2, 3)]:
        ws.cell(row=3, column=14 + offset, value=f"Bank {bank_num}").font = label_font
        ws.cell(row=3, column=19 + offset, value=f"Bank {bank_num}").font = label_font

    # "Largest Min" headers — row 4
    ws.cell(row=4, column=22, value="Largest Min with PEDs").font = label_font
    ws.cell(row=4, column=23, value="Largest Min without PEDs").font = label_font

    # Phase rows 5-12
    for phase_num in range(1, 9):
        row = 4 + phase_num

        # Without PEDs section (cols M-P)
        ws.cell(row=row, column=13, value=f"Phase {phase_num}")
        # With PEDs section (cols R-U)
        ws.cell(row=row, column=18, value=f"Phase {phase_num}")

        max_with_peds = 0.0
        max_without_peds = 0.0

        for bank_num in range(1, 4):
            bank_offset = bank_num - 1
            t = timings_by_bank.get(bank_num, {}).get(phase_num)
            if t:
                without_peds = _compute_min_split_without_peds(t)
                with_peds = _compute_min_split_with_peds(t)
            else:
                without_peds = 0.0
                with_peds = 0.0

            ws.cell(row=row, column=14 + bank_offset, value=without_peds)
            ws.cell(row=row, column=19 + bank_offset, value=with_peds)
            max_without_peds = max(max_without_peds, without_peds)
            max_with_peds = max(max_with_peds, with_peds)

        # Largest columns
        ws.cell(row=row, column=22, value=max_with_peds)
        ws.cell(row=row, column=23, value=max_without_peds)

    # Build lookup: phase -> largest min split (with PEDs) for red highlighting
    largest_min_peds = {}
    for phase_num in range(1, 9):
        max_val = 0.0
        for bank_num in range(1, 4):
            t = timings_by_bank.get(bank_num, {}).get(phase_num)
            if t:
                max_val = max(max_val, _compute_min_split_with_peds(t))
        largest_min_peds[phase_num] = max_val

    # --- Location & Asset (rows 11-12, cols A-B) ---
    ws.cell(row=11, column=1, value="Location:").font = label_font
    ws.cell(row=11, column=2, value=intersection.location_name or "")
    ws.cell(row=12, column=1, value="Asset Number:").font = label_font
    ws.cell(row=12, column=2, value=intersection.asset_number)

    # --- Plan groups: 1-10 (rows 14-26), 11-20 (rows 28-40), 21-30 (rows 42-54) ---
    plan_groups = [
        (1, 10, 14),   # plans 1-10, starting at row 14
        (11, 20, 28),  # plans 11-20, starting at row 28
        (21, 30, 42),  # plans 21-30, starting at row 42
    ]

    for plan_start, plan_end, base_row in plan_groups:
        # Row 0: "Plan" label + plan numbers
        ws.cell(row=base_row, column=1, value="Plan").font = label_font
        for idx, pn in enumerate(range(plan_start, plan_end + 1)):
            col = 2 + idx * 2  # B, D, F, H, J, L, N, P, R, T
            cell = ws.cell(row=base_row, column=col, value=pn)
            cell.font = label_font
            cell.fill = header_fill
            cell.border = thin_border
            # Merge the plan number across 2 columns
            ws.merge_cells(
                start_row=base_row, start_column=col,
                end_row=base_row, end_column=col + 1,
            )

        # Row 1: BiTrans / SEPAC sub-headers
        sub_row = base_row + 1
        for idx in range(10):
            col = 2 + idx * 2
            bt_cell = ws.cell(row=sub_row, column=col, value="BiTrans")
            bt_cell.font = Font(bold=True, size=9)
            bt_cell.fill = header_fill
            bt_cell.border = thin_border
            sp_cell = ws.cell(row=sub_row, column=col + 1, value="SEPAC")
            sp_cell.font = Font(bold=True, size=9, color="006100")
            sp_cell.fill = sepac_fill
            sp_cell.border = thin_border

        # Row 2: Cycle Length
        cl_row = base_row + 2
        ws.cell(row=cl_row, column=1, value="Cycle Length").font = label_font
        for idx, pn in enumerate(range(plan_start, plan_end + 1)):
            col = 2 + idx * 2
            p = plan_map.get(pn)
            cl = p.cycle_length if p else 0.0
            ws.cell(row=cl_row, column=col, value=cl).border = thin_border
            ws.cell(row=cl_row, column=col + 1, value=cl).border = thin_border

        # Rows 3-10: Phase 1-8 ForceOff
        for phase_num in range(1, 9):
            ph_row = base_row + 2 + phase_num
            ws.cell(row=ph_row, column=1, value=f"Phase {phase_num} - ForceOff").font = label_font
            for idx, pn in enumerate(range(plan_start, plan_end + 1)):
                col = 2 + idx * 2
                p = plan_map.get(pn)
                fo = getattr(p, f"phase{phase_num}_force_off", 0.0) if p else 0.0
                split = getattr(p, f"sepac_split{phase_num}", 0.0) if p else 0.0
                split = split or 0.0

                fo_cell = ws.cell(row=ph_row, column=col, value=round(fo, 1))
                fo_cell.border = thin_border
                sp_cell = ws.cell(row=ph_row, column=col + 1, value=round(split, 1))
                sp_cell.border = thin_border

                # Red highlight if SEPAC split < min split (with PEDs)
                min_req = largest_min_peds.get(phase_num, 0.0)
                if split > 0 and min_req > 0 and split < min_req:
                    sp_cell.fill = fail_fill
                    sp_cell.font = fail_font
                else:
                    sp_cell.fill = sepac_fill

        # Row 11: Ring Offset
        ro_row = base_row + 11
        ws.cell(row=ro_row, column=1, value="Ring Offset").font = label_font
        for idx, pn in enumerate(range(plan_start, plan_end + 1)):
            col = 2 + idx * 2
            ws.cell(row=ro_row, column=col, value=0).border = thin_border

        # Row 12: Offset
        off_row = base_row + 12
        ws.cell(row=off_row, column=1, value="Offset").font = label_font
        for idx, pn in enumerate(range(plan_start, plan_end + 1)):
            col = 2 + idx * 2
            p = plan_map.get(pn)
            offset_val = p.offset if p else 0.0
            ws.cell(row=off_row, column=col, value=round(offset_val, 1)).border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, 22):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = 8
    for col_idx in range(22, 24):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # --- Sheet 2: Phase Timing (Bank 1) ---
    ws3 = wb.create_sheet("Phase Timing")
    bank1_timings = sorted(
        [t for t in all_timings if t.bank == 1],
        key=lambda t: t.phase_number,
    )
    timing_headers = ["Phase", "Walk", "FDW", "Min Green", "Veh Ext",
                      "Max 1", "Max 2", "Yellow", "Red Clear"]
    for col, h in enumerate(timing_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = label_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, t in enumerate(bank1_timings, 2):
        ws3.cell(row=row_idx, column=1, value=t.phase_number).border = thin_border
        ws3.cell(row=row_idx, column=2, value=t.ped_walk).border = thin_border
        ws3.cell(row=row_idx, column=3, value=t.ped_fdw).border = thin_border
        ws3.cell(row=row_idx, column=4, value=t.min_green).border = thin_border
        ws3.cell(row=row_idx, column=5, value=t.veh_extension).border = thin_border
        ws3.cell(row=row_idx, column=6, value=t.max_limit_1).border = thin_border
        ws3.cell(row=row_idx, column=7, value=t.max_limit_2).border = thin_border
        ws3.cell(row=row_idx, column=8, value=t.yellow_change).border = thin_border
        ws3.cell(row=row_idx, column=9, value=t.red_clear).border = thin_border

    # Auto-size Phase Timing columns
    for col in ws3.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws3.column_dimensions[col_letter].width = max_len + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf_report(db: Session, asset_number: str) -> io.BytesIO | None:
    """Generate a PDF report for an intersection's conversion results."""
    intersection = db.query(Intersection).filter(
        Intersection.asset_number == asset_number
    ).first()
    if not intersection:
        return None

    timings = db.query(PhaseTiming).filter(
        PhaseTiming.intersection_id == intersection.id,
        PhaseTiming.bank == 1,
    ).order_by(PhaseTiming.phase_number).all()

    plans = db.query(CoordinationPlan).filter(
        CoordinationPlan.intersection_id == intersection.id,
        CoordinationPlan.cycle_length > 0,
    ).order_by(CoordinationPlan.plan_number).all()

    fdot_overrides = db.query(FdotOverride).filter(
        FdotOverride.intersection_id == intersection.id,
    ).order_by(FdotOverride.phase_number, FdotOverride.bank).all()

    min_splits = compute_min_splits(timings, bank=1)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Header
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "BiTrans to SEPAC Conversion Report", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Intersection info
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Intersection Information", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    info_rows = [
        ("Asset Number", asset_number),
        ("Location", intersection.location_name or ""),
        ("Equipment", intersection.equipment_type or ""),
        ("Section", intersection.section or ""),
    ]
    for label, val in info_rows:
        pdf.cell(45, 6, f"{label}:", new_x="RIGHT")
        pdf.cell(0, 6, str(val), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Phase Timing Table (Bank 1)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Phase Timing (Bank 1)", new_x="LMARGIN", new_y="NEXT")
    headers = ["Phase", "Walk", "FDW", "Min Grn", "Yellow", "Red Clr"]
    col_w = 30
    pdf.set_font("Helvetica", "B", 9)
    for h in headers:
        pdf.cell(col_w, 6, h, border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 9)
    for t in timings:
        pdf.cell(col_w, 6, str(t.phase_number), border=1)
        pdf.cell(col_w, 6, str(round(t.ped_walk, 1)), border=1)
        pdf.cell(col_w, 6, str(round(t.ped_fdw, 1)), border=1)
        pdf.cell(col_w, 6, str(round(t.min_green, 1)), border=1)
        pdf.cell(col_w, 6, str(round(t.yellow_change, 1)), border=1)
        pdf.cell(col_w, 6, str(round(t.red_clear, 1)), border=1)
        pdf.ln()
    pdf.ln(4)

    # Coordination Plans with ForceOff + SEPAC Splits
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Coordination Plans - ForceOff & SEPAC Splits", new_x="LMARGIN", new_y="NEXT")

    for plan in plans:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 7, f"Plan {plan.plan_number}  (CL={round(plan.cycle_length, 1)}, Offset={round(plan.offset, 1)})",
                 new_x="LMARGIN", new_y="NEXT")

        pw = 22
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(pw, 5, "", border=1)
        for i in range(1, 9):
            pdf.cell(pw, 5, f"Ph {i}", border=1, align="C")
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        pdf.cell(pw, 5, "ForceOff", border=1)
        for i in range(1, 9):
            val = getattr(plan, f"phase{i}_force_off")
            pdf.cell(pw, 5, str(round(val, 1)) if val else "-", border=1, align="C")
        pdf.ln()

        pdf.cell(pw, 5, "Split", border=1)
        for i in range(1, 9):
            val = getattr(plan, f"sepac_split{i}")
            pdf.cell(pw, 5, str(round(val, 1)) if val else "-", border=1, align="C")
        pdf.ln()

        # Validation row
        sepac = {f"sepac_split{i}": getattr(plan, f"sepac_split{i}") or 0.0 for i in range(1, 9)}
        vals = validate_splits(sepac, min_splits)
        pdf.cell(pw, 5, "Status", border=1)
        for v in vals:
            color = (0, 128, 0) if v["status"] == "pass" else (200, 0, 0) if v["status"] == "fail" else (128, 128, 128)
            pdf.set_text_color(*color)
            pdf.cell(pw, 5, v["status"].upper(), border=1, align="C")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(7)

    # FDOT Overrides Applied
    if fdot_overrides:
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "FDOT Overrides Applied", new_x="LMARGIN", new_y="NEXT")
        ow = [25, 20, 35, 35, 35]
        pdf.set_font("Helvetica", "B", 9)
        for h, w in zip(["Phase", "Bank", "Field", "Original", "FDOT Value"], ow):
            pdf.cell(w, 6, h, border=1)
        pdf.ln()
        pdf.set_font("Helvetica", "", 9)
        for o in fdot_overrides:
            pdf.cell(ow[0], 6, str(o.phase_number), border=1)
            pdf.cell(ow[1], 6, str(o.bank), border=1)
            pdf.cell(ow[2], 6, o.field_name, border=1)
            pdf.cell(ow[3], 6, str(round(o.original_value, 1)), border=1)
            pdf.cell(ow[4], 6, str(round(o.fdot_value, 1)), border=1)
            pdf.ln()

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return output
