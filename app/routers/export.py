"""Export endpoints — Excel and JSON downloads."""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.migration import MigrationRecord
from app.services.export_service import export_intersection_excel, export_intersection_json, generate_pdf_report

router = APIRouter()


@router.get("/{asset}/excel")
def download_excel(asset: str, db: Session = Depends(get_db)):
    output = export_intersection_excel(db, asset)
    if not output:
        raise HTTPException(404, f"Intersection {asset} not found")

    # Update migration status
    migration = db.query(MigrationRecord).filter(
        MigrationRecord.asset_number == asset
    ).first()
    if migration and migration.status == "converted":
        migration.status = "exported"
        db.commit()

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={asset}_SEPAC_conversion.xlsx"},
    )


@router.get("/{asset}/pdf")
def download_pdf(asset: str, db: Session = Depends(get_db)):
    output = generate_pdf_report(db, asset)
    if not output:
        raise HTTPException(404, f"Intersection {asset} not found")

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={asset}_SEPAC_report.pdf"},
    )


@router.get("/{asset}/json")
def download_json(asset: str, db: Session = Depends(get_db)):
    data = export_intersection_json(db, asset)
    if not data:
        raise HTTPException(404, f"Intersection {asset} not found")
    return JSONResponse(content=data)


@router.get("/batch/excel")
def batch_export(
    assets: str = Query(..., description="Comma-separated asset numbers"),
    db: Session = Depends(get_db),
):
    asset_list = [a.strip() for a in assets.split(",") if a.strip()]
    if not asset_list:
        raise HTTPException(400, "No assets specified")

    # For batch, create a multi-sheet workbook
    from openpyxl import Workbook
    from app.models.intersection import Intersection
    from app.models.timing import CoordinationPlan
    import io

    wb = Workbook()
    wb.remove(wb.active)

    for asset in asset_list:
        intersection = db.query(Intersection).filter(
            Intersection.asset_number == asset
        ).first()
        if not intersection:
            continue

        plans = db.query(CoordinationPlan).filter(
            CoordinationPlan.intersection_id == intersection.id,
            CoordinationPlan.cycle_length > 0,
        ).order_by(CoordinationPlan.plan_number).all()

        ws = wb.create_sheet(title=f"Asset {asset}")
        ws.append(["Plan", "CL", "Offset",
                    "Ph1 FO", "Ph1 Split", "Ph2 FO", "Ph2 Split",
                    "Ph3 FO", "Ph3 Split", "Ph4 FO", "Ph4 Split",
                    "Ph5 FO", "Ph5 Split", "Ph6 FO", "Ph6 Split",
                    "Ph7 FO", "Ph7 Split", "Ph8 FO", "Ph8 Split"])

        for p in plans:
            row = [p.plan_number, p.cycle_length, p.offset]
            for i in range(1, 9):
                row.append(getattr(p, f"phase{i}_force_off"))
                row.append(getattr(p, f"sepac_split{i}"))
            ws.append(row)

    if not wb.sheetnames:
        raise HTTPException(404, "No matching intersections found")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=batch_SEPAC_conversion.xlsx"},
    )
