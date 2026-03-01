"""Parse the MDC Data Prep XLSM master list workbook.

Extracts asset_number, polygon, location, and FDOT phase clearance timings
(Yellow, Red Clearance, PED Clear) for each of 8 phases.
"""

import json
import logging
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

# Column layout for the Master Intersection List sheet.
# Col 0 = Polygon, Col 1 = Asset Number.
# Then 8 phase groups follow. Odd phases have 9 cols, even phases have 11 cols
# (even phases include extra MDC Yellow / MDC Red columns).
#
# Within each phase group the offsets from the group start are:
#   0=Speed, 1=Grade, 2=Width, 3=Crosswalk,
#   4=CalcYellow, 5=CalcRed, 6=Yellow, 7=Red,
#   [8=MDC Yellow, 9=MDC Red] (even phases only),
#   last=PED Clear
#
# Absolute column indices for Yellow / Red / PED Clear per phase:
PHASE_FDOT_COLS = {
    1: {"yellow": 8,  "red": 9,  "ped_clear": 10},  # cols I, J, K
    2: {"yellow": 17, "red": 18, "ped_clear": 21},   # cols R, S, V
    3: {"yellow": 28, "red": 29, "ped_clear": 30},   # cols AC, AD, AE
    4: {"yellow": 37, "red": 38, "ped_clear": 41},   # cols AL, AM, AP
    5: {"yellow": 48, "red": 49, "ped_clear": 50},   # cols AW, AX, AY
    6: {"yellow": 57, "red": 58, "ped_clear": 61},   # cols BF, BG, BJ
    7: {"yellow": 68, "red": 69, "ped_clear": 70},   # cols BQ, BR, BS
    8: {"yellow": 77, "red": 78, "ped_clear": 81},   # cols BZ, CA, CD
}

# Header row (0-indexed) in the Master Intersection List sheet.
# Row 3 = phase group labels, Row 4 = column headers, data starts at Row 5.
HEADER_ROW = 4  # 0-indexed (Excel row 5)
DATA_START_ROW = 5  # 0-indexed (Excel row 6)


def _safe_float(val) -> float | None:
    """Convert a cell value to float, returning None for empty/missing."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val == "-":
            return None
        try:
            return float(val)
        except ValueError:
            return None
    return None


def parse_master_list(file_path: str | Path) -> list[dict]:
    """Parse the master intersection list from the XLSM workbook.

    Returns:
        List of dicts with asset_number, polygon, location, phases_json.
        phases_json is a JSON string mapping phase number to
        {yellow, red_clear, ped_clear} values.
    """
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

    # Find the master list sheet
    target_sheet = None
    for name in wb.sheetnames:
        lower = name.lower()
        if "master" in lower and ("list" in lower or "intersection" in lower):
            target_sheet = wb[name]
            break

    if not target_sheet:
        for name in wb.sheetnames:
            lower = name.lower()
            if "master" in lower or "list" in lower:
                target_sheet = wb[name]
                break

    if not target_sheet:
        target_sheet = wb[wb.sheetnames[0]]

    rows = list(target_sheet.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    # Find header row — look for "Asset" or "Number" in first 6 rows
    header_row_idx = 0
    headers = []
    for i, row in enumerate(rows[:6]):
        row_strs = [str(c).lower() if c else "" for c in row]
        if any("asset" in s for s in row_strs):
            header_row_idx = i
            headers = [str(c).strip() if c else "" for c in row]
            break

    if not headers:
        headers = [str(c).strip() if c else "" for c in rows[0]]

    # Identify asset and polygon columns from headers
    asset_col = None
    polygon_col = None
    location_col = None
    for idx, h in enumerate(headers):
        hl = h.lower()
        if "asset" in hl or ("number" in hl and "phone" not in hl):
            asset_col = idx
        elif "polygon" in hl or "poly" in hl:
            polygon_col = idx
        elif "location" in hl or "intersection" in hl:
            location_col = idx

    if asset_col is None:
        # Default: col B (index 1) is typical for this sheet
        asset_col = 1
    if polygon_col is None:
        polygon_col = 0  # col A is typically Polygon

    entries = []
    for row in rows[header_row_idx + 1:]:
        if not row or len(row) <= asset_col:
            continue

        asset_val = row[asset_col]
        if asset_val is None:
            continue

        # Convert to string
        if isinstance(asset_val, float):
            asset_val = str(int(asset_val))
        else:
            asset_val = str(asset_val).strip()

        if not asset_val or asset_val == "0":
            continue

        polygon = ""
        if polygon_col is not None and len(row) > polygon_col and row[polygon_col]:
            pv = row[polygon_col]
            if isinstance(pv, float):
                polygon = str(int(pv))
            else:
                polygon = str(pv).strip()

        location = ""
        if location_col is not None and len(row) > location_col and row[location_col]:
            location = str(row[location_col]).strip()

        # Extract FDOT phase timings
        phases = {}
        for phase_num, cols in PHASE_FDOT_COLS.items():
            yellow = _safe_float(row[cols["yellow"]]) if len(row) > cols["yellow"] else None
            red = _safe_float(row[cols["red"]]) if len(row) > cols["red"] else None
            ped = _safe_float(row[cols["ped_clear"]]) if len(row) > cols["ped_clear"] else None

            # Only include phase if at least one value is present
            if yellow is not None or red is not None or ped is not None:
                phases[str(phase_num)] = {
                    "yellow": yellow or 0.0,
                    "red_clear": red or 0.0,
                    "ped_clear": ped or 0.0,
                }

        phases_json = json.dumps(phases) if phases else None

        entries.append({
            "asset_number": asset_val,
            "polygon": polygon,
            "location": location,
            "phases_json": phases_json,
        })

    wb.close()
    logger.info("Parsed %d entries from master list, %d with FDOT phases",
                len(entries), sum(1 for e in entries if e["phases_json"]))
    return entries
